import numpy as np
import pandas as pd
import random
from datetime import datetime

df = pd.read_excel('../請放入--新的人力預估、進線量預估/任意舊班表.xlsx')
staff_needed = pd.read_excel('../請放入--新的人力預估、進線量預估/人力預估.xlsx')
phone_traffic = pd.read_excel('../請放入--新的人力預估、進線量預估/進線量預估.xlsx')
date = staff_needed.iloc[:, 0]
df.iloc[:, 0] = staff_needed.iloc[:, 0]

# indices of week
dates = np.array(staff_needed['time'])
weekdays = []
weekends = []

for (i, d) in enumerate(dates):
    if datetime.strptime(str(d)[:10], "%Y-%m-%d").weekday() < 5:
        weekdays.append(i)
    else:
        weekends.append(i)
week = []
for i in range(4):
    week.append(weekdays[i * 5:i * 5 + 5])


# print(week) # [[1, 2, 3, 4, 5], [8, 9, 10, 11, 12], [15, 16, 17, 18, 19], [22, 23, 24, 25, 26]]

# 周末 需求轉各時段人力
def demand_transform_weekend(x):
    for i in range(17):
        x[i] = x[i] - 1  # 夜班 0 - 8 共 17 時段 減 1
    x = x[14:]  # 7:00 開始需求（假設前面滿足）
    threshold = 1
    x = np.ceil(x * threshold)
    output = []
    for i in range(20):
        if x[i] < 0:
            x[i] = 0
        elif x[i] == 1:
            x[i] = 2  # 不能落單
        output.append(x[i])
        x[(i + 8):(i + 9)] = x[(i + 8):(i + 9)] + x[i]
        x[i:(17 + i)] = x[i:(17 + i)] - x[i]
    return np.array(output)  # 7:00 - 16:30 共 20 時段


# 周間 需求轉各時段人力
def demand_transform_weekday(x, threshold):
    for i in range(17):
        x[i] = x[i] - 1  # 夜班 0 - 8 共 17 時段 減 1
    x = x[13:]  # 6:30 開始需求（假設前面滿足）
    x[8:10] = x[8:10] + 1
    x[0:18] = x[0:18] - 1  # 扣掉 6:30
    x[10:12] = x[10:12] + 2
    x[2:20] = x[2:20] - 2  # 扣掉 7:30
    x = x[3:]
    x = np.ceil(x * threshold)
    output = []

    for i in range(18):

        if i in [12, 13, 14, 15, 16, 17]:  # 14:00 以後不開會

            if x[i] < 0:
                x[i] = 0
            elif x[i] == 1:
                x[i] = 2
            output.append(x[i])
            x[(i + 8):(i + 9)] = x[(i + 8):(i + 9)] + x[i]
            x[i:(i + 17)] = x[i:(i + 17)] - x[i]

        elif i in range(12):  # [ 0,  1,  2,  3, 11]: # 限定上班時段

            if x[i + 1] < 0:
                x[i + 1] = 0
            elif x[i + 1] == 1:
                x[i + 1] = 2
            output.append(x[i + 1])
            x[(i + 8):(i + 10)] = x[(i + 8):(i + 10)] + x[i + 1]
            x[(i + 1):(i + 18)] = x[(i + 1):(i + 18)] - x[i + 1]

    return output  # 8:00 - 16:30 共 21 時段


def adjust_demands(x, num):
    if sum(x) < num:
        sort_index_x = sorted(range(len(x)), key=lambda k: x[k], reverse=True)
        for i in range(num - sum(x)):
            x[sort_index_x[i % len(sort_index_x)]] += 1
    elif sum(x) > num:
        idx = -1
        while sum(x) > num:
            if x[idx] != 0:
                x[idx] -= 1
            else:
                idx -= 1
    return x


def project_py(score_threshold):
    #### Input
    all_weekday = demand_transform_weekday(staff_needed.iloc[weekdays, :].mean(axis=0),
                                           threshold=1)  # 8:00 - 16:30 共 21 時段
    all_weekend = demand_transform_weekend(staff_needed.iloc[weekends, :].mean(axis=0))  # 7:00 - 16:30 共 20 時段
    #     print("All Weekday:", all_weekday)
    #     print("All Weekend:", all_weekend)
    # men_weekday = [3,5,5,5,5,0,2] # 平日男班（1330~1630）扣除台中 1630*2 -> 總和: 25
    # men_weekend = [2,0,0,2,2,2,3] # 假日男班（1330~1630） -> 總和: 11
    #               [3,1,3,1,0,0,3]
    #               [1,1,3,1,1,1,3]
    # x = [20,17,9,8,4,2,2,24,2,0,0,0] # 平日女班（0800~1330）扣除特殊早班 0630*1 + 0730*2 -> 總和: 88
    #                 [23,14,9,6,4,2,2,21,2,0,0,0]
    #                 [24,15,10,7,4,2,2,22,2,0,0,0]
    # women_weekend = [2,0,2,2,3,2,2,0,3,0,0,2,0,0] # 假日女班（0700~1330） -> 總和: 18
    #                 [2,0,2,2,3,2,2,0,3,0,0,2,0,2]
    #                 [2,0,2,2,3,2,2,0,3,0,0,2,0,0]

    # 平日女班（0800~1330）扣除特殊早班 0630*1 + 0730*2 -> 總和: 88
    women_weekday = [int(wd) for wd in all_weekday[0:12]]
    women_weekday = adjust_demands(women_weekday, 88)
    #     print("Women Weekday:", sum(women_weekday), women_weekday) # Women Weekday: 88 [24, 15, 10, 7, 4, 2, 2, 22, 2, 0, 0, 0]

    # 假日女班（0700~1330） -> 總和: 22
    women_weekend = [int(we) for we in all_weekend[0:14]]
    women_weekend = adjust_demands(women_weekend, 22)
    #     print("Women Weekend:", sum(women_weekend), women_weekend) # Women Weekend: 18 [2, 0, 2, 2, 3, 2, 2, 0, 3, 0, 0, 2, 0, 0]

    # 平日男班（1330~1630）扣除台中 1630*2 -> 總和: 25
    men_weekday = [int(md) for md in all_weekday[12:18]]
    men_weekday = [0] + men_weekday
    men_weekday = adjust_demands(men_weekday, 25)
    #     print("Men Weekday:", sum(men_weekday), men_weekday) # Men Weekday: 25 [3, 2, 2, 2, 12, 2, 2]

    # 假日男班（1330~1630） -> 總和: 11
    men_weekend = [int(me) for me in all_weekend[14:20]]
    men_weekend = [0] + men_weekend
    men_weekend = adjust_demands(men_weekend, 11)
    #     print("Men Weekend:", sum(men_weekend), men_weekend) # Men Weekend: 11 [1, 1, 3, 1, 1, 1, 3]

    ##### 女班
    ### 台中女班
    taichung_girl = list(range(76, 92))
    df[taichung_girl] = 0

    taichung_girl_1 = [78, 79, 80, 83, 87, 88, 89, 91]
    taichung_girl_2 = [76, 77, 81, 82, 84, 85, 86, 90]

    women_weekday_time = ['0800', '0830', '0900', '0930', '1000', '1030', '1100', '1130', '1200', '1230', '1300',
                          '1330']

    for i in range(len(women_weekday) - 1, -1, -1):
        if women_weekday[i] >= 8:
            taichung_afternoon_time = women_weekday_time[i]
            women_weekday[i] -= 8
            break
    women_weekday[0] -= 8
    df.iloc[:14, taichung_girl_1] = '0800'
    df.iloc[14:, taichung_girl_1] = taichung_afternoon_time
    df.iloc[:14, taichung_girl_2] = taichung_afternoon_time
    df.iloc[14:, taichung_girl_2] = '0800'

    df.iloc[weekends, taichung_girl] = '休'

    ### 其他所有女班

    # 特殊班排班 (19,29,34 與 k2 的 5,11,28 交換代號 以 1.維持模型穩定性 2. 增加開會出席率)
    early_bean_1 = [5]
    early_bean_2 = [11, 28]
    early_bean = early_bean_1 + early_bean_2
    df.iloc[:, early_bean_1] = '0630'
    df.iloc[:, early_bean_2] = '0730'
    df.iloc[weekends, early_bean] = '休'

    # 表先全填休 (debug)
    girls = list(range(1, 76))
    girls = [x for x in girls if x not in early_bean]
    df.iloc[:, girls] = '休'

    ## 女平日班

    # indices of group k1, k3, k4, k5, k6, k7
    k = [[6, 12, 14, 17, 25, 40, 48, 54],
         [4, 7, 8, 10, 16, 19, 22, 26],
         [13, 15, 23, 24, 29, 32, 36, 50],
         [1, 2, 9, 31, 35, 42, 47, 55],
         [18, 33, 37, 39, 44, 45, 46, 51],
         [3, 20, 21, 27, 30, 34, 41, 52]]

    # assign starting time to group k1, k3, k4, k5, k6, k7 (same group same starting time)(random assignment -> fair)
    start_t = []
    for i in range(len(women_weekday)):
        if i == 1:  ##### 特例: 0830 只配一組
            start_t += ([women_weekday_time[i]] * 1)
            women_weekday[i] -= 8
            continue
        start_t += ([women_weekday_time[i]] * (women_weekday[i] // 8))
        women_weekday[i] -= (8 * (women_weekday[i] // 8))
    random.shuffle(start_t)
    start_t = start_t * 2
    for i in range(6):
        df.iloc[week[0], k[i]] = start_t[i]
        df.iloc[week[1], k[i]] = start_t[i + 1]
        df.iloc[week[2], k[i]] = start_t[i + 2]
        df.iloc[week[3], k[i]] = start_t[i + 3]

    # assign starting time to group k2, t1, t2, t3 (random starting time across groups)
    start_t_rd = []
    for i in range(len(women_weekday)):
        start_t_rd += ([women_weekday_time[i]] * women_weekday[i])
    random.shuffle(start_t_rd)
    rest_girls = [38, 43, 49, 53,  # k2
                  56, 58, 61, 65, 67, 70,  # t1
                  62, 63, 64, 68, 69, 73, 74,  # t2
                  57, 59, 60, 66, 71, 72, 75]  # t3
    start_t_rd = start_t_rd * 2
    for i in range(24):  # 24 people
        df.iloc[week[0], rest_girls[i]] = start_t_rd[i]
        df.iloc[week[1], rest_girls[i]] = start_t_rd[i + 1]
        df.iloc[week[2], rest_girls[i]] = start_t_rd[i + 2]
        df.iloc[week[3], rest_girls[i]] = start_t_rd[i + 3]

    ## 女假日班
    # women_weekend = [2,0,2,2,3,2,2,0,3,0,0,2,0,0] # input [2,0,0,2,2,3,2,2,0,3,0,0,2,0,0]

    taichung_girl = list(range(76, 92)) * 2
    taipei_girl = list(range(56, 76)) * 2
    kaohsiung_girl = [x for x in list(range(1, 56)) if x not in early_bean] * 2
    weekend_time = ['0630', '0700', '0730', '0800', '0830', '0900', '0930', '1000', '1030', '1100', '1130', '1200',
                    '1230', '1300', '1330']
    weekend_time_ls = []
    for i in range(len(women_weekend)):
        weekend_time_ls += (women_weekend[i] * [weekend_time[i]])
    tp_tc = taichung_girl[:10] + taipei_girl[:12] + taichung_girl[10:20] + taipei_girl[12:24] + \
            taichung_girl[20:30] + taipei_girl[24:36] + taichung_girl[30:] + taipei_girl[36:]
    ks = kaohsiung_girl

    df.iloc[weekends[0], tp_tc[:22]] = weekend_time_ls
    df.iloc[weekends[3], tp_tc[22:44]] = weekend_time_ls
    df.iloc[weekends[4], tp_tc[44:66]] = weekend_time_ls
    df.iloc[weekends[2], tp_tc[66:]] = weekend_time_ls[:6]
    df.iloc[weekends[2], ks[:16]] = weekend_time_ls[6:]
    df.iloc[weekends[1], ks[16:38]] = weekend_time_ls
    df.iloc[weekends[5], ks[38:60]] = weekend_time_ls
    df.iloc[weekends[6], ks[60:82]] = weekend_time_ls
    df.iloc[weekends[7], ks[82:]] = weekend_time_ls


    # 女平日休假 (固定二三四五，每八人休一人)

    ## k_group (k1,k3,k4,k5,k6,k7)
    days_off = week[0][1:] + week[1][1:] + week[2][1:] + week[3][1:]
    mod_k = np.array(k)
    mod_k = np.hstack((k, k))

    ## rest
    random.shuffle(rest_girls)
    rg = np.array([rest_girls[:8], rest_girls[8:16], rest_girls[16:]])
    rg = np.hstack((rg, rg))

    # tc
    tc1 = taichung_girl_1
    tc2 = taichung_girl_2
    tc = np.array([tc1 * 2, tc2 * 2])

    for i in range(16):
        df.iloc[days_off[i], mod_k[:, i]] = '休'
        df.iloc[days_off[i], rg[:, i]] = '休'
        df.iloc[days_off[i], tc[:, i]] = '休'

    # 休差
    for j in girls:
        i_1 = -1
        i_2_exist = False
        for i in range(28):

            if (df.iloc[i, j] == '休') & (not i_2_exist):

                i_2_exist = True
                i_2 = i

                if i_2 - i_1 <= 7:
                    pass
                elif i_2 - i_1 == 8:
                    i_minus = random.randint(2, 6)
                    df.iloc[i - i_minus, j] = '休'

            if (df.iloc[i, j] == '休') & i_2_exist:

                i_1 = i_2
                i_2 = i

                if i_2 - i_1 <= 7:
                    pass
                elif i_2 - i_1 == 8:
                    i_minus = random.randint(2, 6)
                    df.iloc[i - i_minus, j] = '休'

            if (i == 27) and (df.iloc[i, j] != '休'):

                i_1 = i_2
                i_2 = 28

                if i_2 - i_1 <= 7:
                    pass
                elif i_2 - i_1 == 8:
                    i_minus = random.randint(2, 6)
                    df.iloc[i - i_minus, j] = '休'

    ##### 男班
    all_men = list(range(92, 119))
    taipei_kao = list(range(92, 117))
    df[all_men] = 0
    kaohsiung_men = []
    for i in range(92, 98):
        kaohsiung_men.append(i)
    for i in range(100, 110):
        kaohsiung_men.append(i)
    taipei_men = [98, 99, 110, 111, 112, 113, 114, 115, 116]
    taichung_men = [117, 118]

    # 男假日班
    # men_weekend = [2,0,0,2,2,2,3] # 假日男班（1330~1630）
    weekend_time = ['1330', '1400', '1430', '1500', '1530', '1600', '1630']
    weekend_time_ls = []
    for i in range(len(men_weekend)):
        weekend_time_ls += (men_weekend[i] * [weekend_time[i]])

    tp_tc = (taipei_men + taichung_men) * 3
    ks = kaohsiung_men * 3 + kaohsiung_men[:7]

    df.iloc[weekends[0], tp_tc[:11]] = weekend_time_ls
    df.iloc[weekends[3], tp_tc[11:22]] = weekend_time_ls
    df.iloc[weekends[4], tp_tc[22:33]] = weekend_time_ls
    df.iloc[weekends[2], ks[:11]] = weekend_time_ls
    df.iloc[weekends[1], ks[11:22]] = weekend_time_ls
    df.iloc[weekends[5], ks[22:33]] = weekend_time_ls
    df.iloc[weekends[6], ks[33:44]] = weekend_time_ls
    df.iloc[weekends[7], ks[44:]] = weekend_time_ls

    # 男大夜班
    tp_tc = (taipei_men + taichung_men)
    ks = kaohsiung_men

    df.iloc[0, 98] = '0030'
    for i, x in enumerate(range(7, 17)):
        df.iloc[x, tp_tc[1:][i]] = '0030'
        df.iloc[x - 1, tp_tc[1:][i]] = '休'
    for i, x in enumerate(range(1, 7)):
        df.iloc[x, ks[:6][i]] = '0030'
        df.iloc[x - 1, ks[:6][i]] = '休'
    for i, x in enumerate(range(17, 27)):
        df.iloc[x, ks[6:][i]] = '0030'
        df.iloc[x - 1, ks[6:][i]] = '休'
    df.iloc[27, 92] = '0030'
    df.iloc[26, 92] = '休'

    # 假日休
    for i in range(len(weekends)):
        for j in range(92, 119):
            if df.iloc[weekends[i], j] == 0:
                df.iloc[weekends[i], j] = "休"

    # 休差
    for j in all_men:
        i_1 = -1
        i_2_exist = False
        for i in range(28):

            if (df.iloc[i, j] == '休') & (not i_2_exist):

                i_2_exist = True
                i_2 = i

                if i_2 - i_1 <= 6:
                    pass
                elif i_2 - i_1 == 7:
                    i_minus = random.randint(1, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(1, 6)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 8:
                    i_minus = random.randint(2, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(2, 6)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 9:
                    i_minus = random.randint(3, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(3, 6)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 10:
                    i_minus = random.randint(4, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(4, 6)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 11:
                    if df.iloc[i - 5, j] == 0:
                        df.iloc[i - 5, j] = '休'
                    elif df.iloc[i - 6, j] == 0:
                        df.iloc[i - 6, j] = '休'
                    else:
                        i_minus = random.randint(3, 4)
                        while (df.iloc[i - i_minus, j] != 0):
                            i_minus = random.randint(3, 4)
                        df.iloc[i - i_minus, j] = '休'
                        i_minus = random.randint(7, 9)
                        while (df.iloc[i - i_minus, j] != 0):
                            i_minus = random.randint(7, 9)
                        df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 12:
                    if df.iloc[i - 6, j] == 0:
                        df.iloc[i - 6, j] = '休'
                    else:
                        i_minus = random.randint(3, 5)
                        while (df.iloc[i - i_minus, j] != 0):
                            i_minus = random.randint(3, 5)
                        df.iloc[i - i_minus, j] = '休'
                        i_minus = random.randint(7, 9)
                        while (df.iloc[i - i_minus, j] != 0):
                            i_minus = random.randint(7, 9)
                        df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 13:
                    i_minus = random.randint(4, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(4, 6)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(8, 10)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(8, 10)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 14:
                    i_minus = random.randint(4, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(4, 6)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(8, 10)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(8, 10)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 15:
                    i_minus = random.randint(2, 4)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(2, 4)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(6, 8)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(6, 8)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(10, 12)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(10, 12)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 16:
                    i_minus = random.randint(2, 4)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(2, 4)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(6, 8)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(6, 8)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(10, 12)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(10, 12)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 19:
                    df.iloc[i - 4, j] = '休'
                    df.iloc[i - 10, j] = '休'
                    i_minus = random.randint(14, 16)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 20:
                    df.iloc[i - 4, j] = '休'
                    df.iloc[i - 10, j] = '休'
                    i_minus = random.randint(15, 16)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 21:
                    df.iloc[i - 5, j] = '休'
                    df.iloc[i - 11, j] = '休'
                    i_minus = random.randint(15, 17)
                    df.iloc[i - i_minus, j] = '休'

            if (df.iloc[i, j] == '休') & i_2_exist:

                i_1 = i_2
                i_2 = i

                if i_2 - i_1 <= 6:
                    pass
                elif i_2 - i_1 == 7:
                    i_minus = random.randint(1, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(1, 6)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 8:
                    i_minus = random.randint(2, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(2, 6)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 9:
                    i_minus = random.randint(3, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(3, 6)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 10:
                    i_minus = random.randint(4, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(4, 6)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 11:
                    if df.iloc[i - 5, j] == 0:
                        df.iloc[i - 5, j] = '休'
                    elif df.iloc[i - 6, j] == 0:
                        df.iloc[i - 6, j] = '休'
                    else:
                        i_minus = random.randint(3, 4)
                        while (df.iloc[i - i_minus, j] != 0):
                            i_minus = random.randint(3, 4)
                        df.iloc[i - i_minus, j] = '休'
                        i_minus = random.randint(7, 9)
                        while (df.iloc[i - i_minus, j] != 0):
                            i_minus = random.randint(7, 9)
                        df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 12:
                    if df.iloc[i - 6, j] == 0:
                        df.iloc[i - 6, j] = '休'
                    else:
                        i_minus = random.randint(3, 5)
                        while (df.iloc[i - i_minus, j] != 0):
                            i_minus = random.randint(3, 5)
                        df.iloc[i - i_minus, j] = '休'
                        i_minus = random.randint(7, 9)
                        while (df.iloc[i - i_minus, j] != 0):
                            i_minus = random.randint(7, 9)
                        df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 13:
                    i_minus = random.randint(4, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(4, 6)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(8, 10)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(8, 10)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 14:
                    i_minus = random.randint(4, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(4, 6)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(8, 10)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(8, 10)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 15:
                    i_minus = random.randint(2, 4)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(2, 4)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(6, 8)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(6, 8)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(10, 12)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(10, 12)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 16:
                    i_minus = random.randint(2, 4)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(2, 4)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(6, 8)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(6, 8)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(10, 12)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(10, 12)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 19:
                    df.iloc[i - 4, j] = '休'
                    df.iloc[i - 10, j] = '休'
                    i_minus = random.randint(14, 16)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 20:
                    df.iloc[i - 4, j] = '休'
                    df.iloc[i - 10, j] = '休'
                    i_minus = random.randint(15, 16)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 21:
                    df.iloc[i - 5, j] = '休'
                    df.iloc[i - 11, j] = '休'
                    i_minus = random.randint(15, 17)
                    df.iloc[i - i_minus, j] = '休'

            if (i == 27) and (df.iloc[i, j] != '休'):

                i_1 = i_2
                i_2 = 28

                if i_2 - i_1 <= 6:
                    pass
                elif i_2 - i_1 == 7:
                    i_minus = random.randint(1, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(1, 6)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 8:
                    i_minus = random.randint(2, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(2, 6)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 9:
                    i_minus = random.randint(3, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(3, 6)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 10:
                    i_minus = random.randint(4, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(4, 6)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 11:
                    if df.iloc[i - 5, j] == 0:
                        df.iloc[i - 5, j] = '休'
                    elif df.iloc[i - 6, j] == 0:
                        df.iloc[i - 6, j] = '休'
                    else:
                        i_minus = random.randint(3, 4)
                        while (df.iloc[i - i_minus, j] != 0):
                            i_minus = random.randint(3, 4)
                        df.iloc[i - i_minus, j] = '休'
                        i_minus = random.randint(7, 9)
                        while (df.iloc[i - i_minus, j] != 0):
                            i_minus = random.randint(7, 9)
                        df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 12:
                    if df.iloc[i - 6, j] == 0:
                        df.iloc[i - 6, j] = '休'
                    else:
                        i_minus = random.randint(3, 5)
                        while (df.iloc[i - i_minus, j] != 0):
                            i_minus = random.randint(3, 5)
                        df.iloc[i - i_minus, j] = '休'
                        i_minus = random.randint(7, 9)
                        while (df.iloc[i - i_minus, j] != 0):
                            i_minus = random.randint(7, 9)
                        df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 13:
                    i_minus = random.randint(4, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(4, 6)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(8, 10)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(8, 10)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 14:
                    i_minus = random.randint(4, 6)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(4, 6)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(8, 10)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(8, 10)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 15:
                    i_minus = random.randint(2, 4)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(2, 4)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(6, 8)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(6, 8)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(10, 12)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(10, 12)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 16:
                    i_minus = random.randint(2, 4)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(2, 4)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(6, 8)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(6, 8)
                    df.iloc[i - i_minus, j] = '休'
                    i_minus = random.randint(10, 12)
                    while (df.iloc[i - i_minus, j] != 0):
                        i_minus = random.randint(10, 12)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 19:
                    df.iloc[i - 4, j] = '休'
                    df.iloc[i - 10, j] = '休'
                    i_minus = random.randint(14, 16)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 20:
                    df.iloc[i - 4, j] = '休'
                    df.iloc[i - 10, j] = '休'
                    i_minus = random.randint(15, 16)
                    df.iloc[i - i_minus, j] = '休'
                elif i_2 - i_1 == 21:
                    df.iloc[i - 5, j] = '休'
                    df.iloc[i - 11, j] = '休'
                    i_minus = random.randint(15, 17)
                    df.iloc[i - i_minus, j] = '休'

    ## 補足八天假
    rest_by_person = (df.iloc[:, all_men] == '休').sum(axis=0).values
    rest_by_person[rest_by_person > 8] = 8
    rest_n_diff = 8 * 25 - rest_by_person.sum()

    rest_daily = (df.iloc[:, all_men] == '休').sum(axis=1).values
    available_rest = []
    # 補到單日四人休
    for i in range(28):
        if rest_daily[i] == 1:
            available_rest += (2 * [i])
        elif rest_daily[i] == 2:
            available_rest += (1 * [i])
    for i in range(28):
        if rest_daily[i] == 1:
            available_rest += (1 * [i])
        elif rest_daily[i] == 2:
            available_rest += (1 * [i])
        elif rest_daily[i] == 3:
            available_rest += (1 * [i])
    # 若不夠，補到單日五人休
    if len(available_rest) < rest_n_diff:
        countdown = rest_n_diff - len(available_rest)
        for i in range(28):
            if rest_daily[i] == 1:
                countdown -= 1
                available_rest += [i]
            elif rest_daily[i] == 2:
                countdown -= 1
                available_rest += [i]
            elif rest_daily[i] == 3:
                countdown -= 1
                available_rest += [i]
            elif rest_daily[i] == 4:
                countdown -= 1
                available_rest += [i]
            if countdown == 0:
                break
    else:
        available_rest = available_rest[:rest_n_diff]

    rest_by_person = (df.iloc[:, all_men] == '休').sum(axis=0).values
    rest_to_go = []
    for i in range(25):
        if rest_by_person[i] < 8:
            rest_to_go += ([92 + i] * (8 - rest_by_person[i]))
    random.shuffle(rest_to_go)

    rest_dic = {}
    for x in available_rest:
        if x not in rest_dic:
            rest_dic[x] = 1
        else:
            rest_dic[x] += 1
    rest_storage = []

    for i, x in enumerate(rest_to_go):

        keys = list(rest_dic.keys())
        find_match = False

        for key in keys:
            if (rest_dic[key] != 0) & (df.iloc[key, x] == 0):
                rest_dic[key] -= 1
                find_match = True
                df.iloc[key, x] = '休'
                rest_storage.append(key)
                break


    # 台中都排1630
    for i in range(len(weekdays)):
        for j in range(len(taichung_men)):
            if df.iloc[weekdays[i], taichung_men[j]] == 0:
                df.iloc[weekdays[i], taichung_men[j]] = '1630'

    # men_weekday = [3,5,5,5,5,0,2] # 平日男班（1330~1630）
    men_weekday_time = ['1330', '1400', '1430', '1500', '1530', '1600', '1630']
    men_ls = list(range(92, 117))
    random.shuffle(men_ls)
    men_group = []  # create a list of groups
    index_sum_1 = 0
    index_sum_2 = 0
    for i in range(len(men_weekday)):
        index_sum_2 += men_weekday[i]
        if i == 0:
            men_group += [men_ls[:men_weekday[i]]]
        else:
            men_group += [men_ls[index_sum_1:index_sum_2]]
        index_sum_1 += men_weekday[i]

    # fill in blank
    for i in range(len(men_group)):
        for y in men_group[i]:
            for j in range(20):
                if (df.iloc[weekdays[j], y] == 0):
                    df.iloc[weekdays[j], y] = men_weekday_time[i]

    ##### meeting time
    # indices of group k1, k2, k3, k4, k5, k6, k7, t1, t2, t3, c1, c2
    group = [[6, 12, 14, 17, 25, 40, 48, 54, 95, 96, 100],
             [5, 11, 28, 38, 43, 49, 53, 102, 106, 107],
             [4, 7, 8, 10, 16, 19, 22, 26, 94, 105],
             [13, 15, 23, 24, 29, 32, 36, 50, 93, 103],
             [1, 2, 9, 31, 35, 42, 47, 55, 92, 97],
             [18, 33, 37, 39, 44, 45, 46, 51, 101, 104],
             [3, 20, 21, 27, 30, 34, 41, 52, 108, 109],
             [56, 58, 61, 65, 67, 70, 111, 112, 113],
             [62, 63, 64, 68, 69, 73, 74, 114, 115, 116],
             [57, 59, 60, 66, 71, 72, 75, 98, 99, 110],
             [78, 79, 80, 83, 87, 88, 89, 91, 118],
             [76, 77, 81, 82, 84, 85, 86, 90, 117]]

    meeting_time = np.empty((28, len(group)), dtype='object')
    meeting_max = np.zeros((28, len(group)))
    meeting_act = np.zeros((28, len(group)))
    meeting_percentage = np.zeros((28, len(group)))
    for i in range(28):
        if i in weekdays:
            for j in range(len(group)):
                inner_group = list(df.iloc[i, group[j]])
                meeting_max[i, j] = sum([(len(x) == 4) for x in inner_group])
                inner_group = [x for x in inner_group if (len(x) == 4)]
                if j in [0, 2, 3, 4, 5, 6, 10, 11]:
                    meeting_time[i, j] = inner_group[0]
                elif j == 1:
                    meeting_time[i, j] = '1400'
                else:
                    meeting_time[i, j] = '1700'
    meeting_act = np.ceil(meeting_max * 0.7)
    meeting_percentage[weekdays, :] = meeting_act[weekdays, :] / meeting_max[weekdays, :]

    ##### evaluate performance
    thirty_min_break = ['1330', '1400', '1430', '1500', '1530', '1600', '1630']

    shifts_arr = df.values[:, 1:]  # pd to np.array
    actual = np.zeros(2 * 24 * 28)

    for i in range(28):
        shifts = shifts_arr[i]
        shifts = shifts[[(len(str(x)) == 4) for x in shifts]]  # filter out nan and those on leave
        shifts_cont = [int(x[:2]) * 2 + int(x[-2:]) // 30 + i * 48 for x in shifts]
        for j in range(len(shifts)):
            x = shifts_cont[j]
            if x > (2 * 24 * 28 - 16):
                actual[x:] += 1
                actual[(x + 2 * 4):(x + 2 * 4 + 1)] -= 1
            elif shifts[j] == '0030':
                actual[x:(x + 2 * 8)] += 1
                actual[(x + 2 * 4):(x + 2 * 4 + 1)] -= 1
            elif i in weekends:
                actual[x:(x + 17)] += 1
                actual[(x + 2 * 4):(x + 2 * 4 + 1)] -= 1
            elif i in weekdays:
                if shifts[j] in thirty_min_break:
                    actual[x:(x + 17)] += 1
                    actual[(x + 2 * 4):(x + 2 * 4 + 1)] -= 1
                else:
                    actual[x:(x + 2 * 9)] += 1
                    actual[(x + 2 * 4):(x + 2 * 4 + 2)] -= 1

    for i in range(28):
        if i in weekdays:
            meetings = meeting_time[i]
            # meetings = meetings[meetings != None]
            meetings = meetings[[(len(str(x)) == 4) for x in meetings]]
            meetings = [int(x[:2]) * 2 + int(x[-2:]) // 30 + i * 48 for x in meetings]
            for j in range(len(meetings)):
                actual[meetings[j]:(meetings[j] + 1)] -= meeting_act[i, j]

    need = staff_needed.values[:, 1:]
    need = need.reshape(need.size, )
    phone = phone_traffic.values[:, 1:]
    phone = phone.reshape(phone.size, )

    w = np.zeros(1344)
    for i in range(1344):
        if 1 <= i % 48 <= 16:
            w[i] = 0.2
        elif 17 <= i % 48 <= 36:
            w[i] = 1
        else:
            w[i] = 0.7

    giveup_rate = 0.862 + 0.05893 * phone + 0.09227 * actual
    service_quality = 70.5838 - 0.4091 * phone + 1.3143 * actual
    service_quality[service_quality > 100] = 100

    total_score = (20 * 1 + 12 * 0.7 + 16 * 0.2) * 28 * 100
    score_1 = (w.T @ (100 - giveup_rate)) / total_score * 100

    score_2 = (w.T @ service_quality) / total_score * 100

    tmp = pd.concat([pd.Series(actual), pd.Series(need)], axis=1)
    tmp.rename(columns={0: 'actual', 1: 'need'}, inplace=True)
    tmp['predicted_call'] = tmp.need * 2.5
    tmp['received_call'] = tmp.apply(lambda row:
                                     row.need * 2.5 if (row.actual > row.need) else
                                     row.actual * 2.5, axis=1
                                     )
    score_3 = (w.T @ tmp.received_call) / (w.T @ tmp.predicted_call) * 100

    tmp['reach_goal'] = (tmp.actual >= tmp.need * 0.7)
    score_4 = tmp.reach_goal.sum() / 1344 * 100

    meeting_time_arr = meeting_time[weekdays, :]
    score_5 = (~np.isnan(meeting_time_arr.astype(np.float))).sum() / 240 * 100

    overall_score = (score_1 * 0.05 + score_2 * 0.05 + score_3 * 0.10 + score_4 * 0.20 + score_5 * 0.20) / (
                0.05 + 0.05 + 0.10 + 0.20 + 0.20)

    if overall_score > score_threshold:
        best_shift_table = df
        best_meeting_percentage = pd.concat([date, pd.DataFrame(np.around(meeting_percentage, decimals=2),
                                                                columns=['k1', 'k2', 'k3', 'k4', 'k5', 'k6', 'k7', 't1',
                                                                         't2', 't3', 'c1', 'c2'])
                                             ], axis=1)
        best_meeting_time = pd.concat([date, pd.DataFrame(meeting_time,
                                                          columns=['k1', 'k2', 'k3', 'k4', 'k5', 'k6', 'k7', 't1', 't2',
                                                                   't3', 'c1', 'c2'])
                                       ], axis=1)
        best_meeting_act = pd.concat([date, pd.DataFrame(meeting_act,
                                                         columns=['k1', 'k2', 'k3', 'k4', 'k5', 'k6', 'k7', 't1', 't2',
                                                                  't3', 'c1', 'c2'])
                                      ], axis=1)

        best_shift_table.to_excel('../情境4班表/新班表_黑白.xlsx')
        best_meeting_percentage.to_excel('../情境4班表/各組開會出席率.xlsx')
        best_meeting_time.to_excel('../情境4班表/各組開會時間.xlsx')
        best_meeting_act.to_excel('../情境4班表/各組開會出席人數.xlsx')

        score_threshold = overall_score

    # return [score_1, score_2, score_3, score_4, score_5, overall_score]
    return [overall_score, score_threshold]


score_threshold = 80  # minimum requirement

for i in range(10):
    tmp, score_threshold = project_py(score_threshold)


#### 彩色班表
from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors, Font, Color, PatternFill


def coloring(path_to_xlsx_file):
    shift_table_read = load_workbook(path_to_xlsx_file)
    shift_table = shift_table_read.active
    row_max = shift_table.max_row
    col_max = shift_table.max_column
    weekend = [2, 8, 9, 15, 16, 22, 23, 29]
    weekday = [i for i in range(2, 30) if i not in weekend]
    color_match = {'0630': 40, '0730': 35, '0800': 26, '0830': 48,
                   '0900': 46, '0930': 44, '1000': 43, '1030': 42,
                   '1100': 41, '1130': 31, '1200': 52, '1330': 22,
                   '1400': 51, '1430': 47, '1500': 55, '1530': 50,
                   '1600': 53, '1630': 33, '1700': 45, '0030': 29, 'weekend': 5}

    for nrow in range(1, row_max + 1):
        if nrow in weekday:
            for ncol in range(3, col_max + 1):
                cell_content = shift_table.cell(row=nrow, column=ncol).value
                if cell_content in color_match:
                    color_ind = color_match[cell_content]
                    background_fill = PatternFill(start_color=colors.COLOR_INDEX[color_ind],
                                                  end_color=colors.COLOR_INDEX[color_ind],
                                                  fill_type='solid')
                    shift_table.cell(row=nrow, column=ncol).fill = background_fill
        elif nrow in weekend:
            for ncol in range(3, col_max + 1):
                cell_content = shift_table.cell(row=nrow, column=ncol).value
                if cell_content != '休':
                    color_ind = color_match['weekend']
                    background_fill = PatternFill(start_color=colors.COLOR_INDEX[color_ind],
                                                  end_color=colors.COLOR_INDEX[color_ind],
                                                  fill_type='solid')
                    shift_table.cell(row=nrow, column=ncol).fill = background_fill

    shift_table_read.save(path_to_xlsx_file[:-8] + '_彩色.xlsx')


coloring('../情境4班表/新班表_黑白.xlsx')